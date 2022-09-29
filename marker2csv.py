from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import glob
import os
import xml.dom.minidom
import numpy as np
import pandas as pd
import utm
import math


def marker2csv(markerFile, csvFile, excelFile, dirname):
    file = xml.dom.minidom.parse(markerFile)
    #print(file.nodeName)
    #print(file.firstChild.tagName)

    Lon = file.getElementsByTagName("LongitudeDegrees")
    Lat = file.getElementsByTagName("LatitudeDegrees")
    Name = file.getElementsByTagName("Name")
    Time = file.getElementsByTagName("RealTime")
    Ping = file.getElementsByTagName("SonarElementIndex")
    File = file.getElementsByTagName("Filename")
    Channel = file.getElementsByTagName("SonarChannel")
    Desc = file.getElementsByTagName("Description")

    LON = []
    LAT = []
    NAME = []
    TIME = []
    PING = []
    FILE = []
    CHANNEL = []
    DESC = []

    LON = [ np.append(LON,float(el.firstChild.nodeValue)) for el in Lon]
    LAT = [ np.append(LAT,float(el.firstChild.nodeValue)) for el in Lat]
    NAME = [ np.append(NAME,el.firstChild.nodeValue) for el in Name]
    TIME = [ np.append(TIME,float(el.firstChild.nodeValue)) for el in Time]
    PING = [ np.append(PING,int(el.firstChild.nodeValue)) for el in Ping]
    FILE = [ np.append(FILE,el.firstChild.nodeValue) for el in File]
    CHANNEL = [ np.append(CHANNEL,int(el.firstChild.nodeValue)) for el in Channel]
    DESC = [ np.append(DESC,el.firstChild.nodeValue) for el in Desc]

    df = np.c_[NAME,LON,LAT,TIME,PING,FILE,CHANNEL,DESC]
    df = pd.DataFrame(df)

    # convert to openpyxl rows
    df_xls = dataframe_to_rows(df)

    # create a workbook and grab active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # images list
    images = []
    for filename in sorted(glob.glob("./{}*.gif".format(dirname))):
        images.append(filename)
        print(filename)

    # resize cells
    for row, im in enumerate(images):
        for col in range(10,11):
            worksheet.row_dimensions[row].height = 300
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 50


    # insert images
    for index, image in enumerate(images):
        worksheet.add_image(Image(image), anchor='J'+str(index+3))
        #print(index,image)


    # insert df
    if True:
        for r_idx, row in enumerate(df_xls, 1):
            for c_idx, value in enumerate(row, 1):
                 worksheet.cell(row=r_idx, column=c_idx, value=value)


    # titles list
    titles = []
    for title in sorted(glob.glob("./{}*.gif".format(dirname))):
        titles.append(os.path.basename(title))


    # save workbook
    workbook.save(dirname + '.xlsx')

